import pandas as pd
from openpyxl import load_workbook

def add_customer_id_for_duplicates(file_path, output_file, start_id=1001):
    # Läs Excel-filen
    df = pd.read_excel(file_path)

    # Kontrollera att nödvändiga kolumner finns
    required_columns = ['First Name', 'Last Name', 'Birthdate', 'Phone']
    if not all(col in df.columns for col in required_columns):
        raise ValueError("Filen saknar en eller flera kolumner: 'First Name', 'Last Name', 'Birthdate', 'Phone'.")

    # Skapa en ny kolumn med unika CustomerID för varje kombination
    unique_combinations = df[['First Name', 'Last Name', 'Birthdate', 'Phone']].drop_duplicates().reset_index(drop=True)
    unique_combinations['CustomerID'] = range(start_id, start_id + len(unique_combinations))

    # Koppla tillbaka CustomerID till original-DataFrame
    df = df.merge(unique_combinations, on=['First Name', 'Last Name', 'Birthdate', 'Phone'], how='left')

    # Spara tillbaka till en ny Excel-fil
    df.to_excel(output_file, index=False, sheet_name="Sheet1")

    # Formatera CustomerID till heltal utan tusentalsseparatorer
    workbook = load_workbook(output_file)
    worksheet = workbook["Sheet1"]

    # Formatera CustomerID (sista kolumnen) till heltal utan tusentalsseparatorer
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                   min_col=worksheet.max_column, max_col=worksheet.max_column):
        for cell in row:
            cell.number_format = '0'
    
    workbook.save(output_file)
    print(f"Uppdaterad fil sparad som '{output_file}' med unika CustomerID baserat på flera kolumner.")

# Använd funktionen
file_path = "customer_data.xlsx"  # Din filväg
output_file = "customer_data_with_customer_id.xlsx"
add_customer_id_for_duplicates(file_path, output_file)
