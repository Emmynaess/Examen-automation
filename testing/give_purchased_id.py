import pandas as pd
from openpyxl import load_workbook

def add_purchase_id(input_file, output_file, purchase_start_id=2001):
    # Läs Excel-filen
    df = pd.read_excel(input_file)

    # Kontrollera att CustomerID finns
    if 'CustomerID' not in df.columns:
        raise ValueError("Filen saknar kolumnen 'CustomerID'. Kontrollera att CustomerID är skapade.")

    # Skapa ett unikt PurchaseID för varje rad
    df['PurchaseID'] = range(purchase_start_id, purchase_start_id + len(df))

    # Spara tillbaka till en ny Excel-fil
    df.to_excel(output_file, index=False, sheet_name="Sheet1")

    # Formatera PurchaseID som heltal utan tusentalsseparatorer
    workbook = load_workbook(output_file)
    worksheet = workbook["Sheet1"]
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                   min_col=worksheet.max_column, max_col=worksheet.max_column):
        for cell in row:
            cell.number_format = '0'
    workbook.save(output_file)

    print(f"Uppdaterad fil sparad som '{output_file}' med unika PurchaseID.")

# Använd funktionen
if __name__ == "__main__":
    input_file = "customer_data_with_customer_id.xlsx"  # Filen med CustomerID
    output_file = "customer_data_with_customer_id_and_purchase_id.xlsx"  # Ny fil med PurchaseID
    add_purchase_id(input_file, output_file)